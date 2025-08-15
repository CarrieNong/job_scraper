import sqlite3
from openpyxl import Workbook, load_workbook
import os

DB_NAME = "jobs.db"
EXCEL_FILE = "jobs_export.xlsx"

def export_jobs_to_excel():
    conn = sqlite3.connect(DB_NAME)
    cursor = conn.cursor()

    # 获取符合条件且未导出的岗位
    cursor.execute("""
        SELECT created_at,title, company, location, applicants, link, job_id
        FROM jobs
        WHERE is_match=1 AND (exported IS NULL OR exported=0)
    """)
    rows = cursor.fetchall()

    if not rows:
        print("没有新岗位可导出")
        conn.close()
        return

    # 如果文件存在则加载，否则新建
    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        # 写入表头
        ws.append(['Created_at', 'Title','Company','Location','Applicants','Link','Job_ID'])

    # 写入岗位数据
    for row in rows:
        ws.append(row)

    wb.save(EXCEL_FILE)
    print(f"成功导出 {len(rows)} 个岗位到 {EXCEL_FILE}")

    # 标记已导出
    # 检查列是否存在
    cursor.execute("PRAGMA table_info(jobs)")
    columns = [info[1] for info in cursor.fetchall()]

    if "exported" not in columns:
        cursor.execute("ALTER TABLE jobs ADD COLUMN exported INTEGER DEFAULT 0")
    cursor.execute("UPDATE jobs SET exported=1 WHERE is_match=1 AND (exported IS NULL OR exported=0)")
    conn.commit()
    conn.close()

if __name__ == "__main__":
    export_jobs_to_excel()
